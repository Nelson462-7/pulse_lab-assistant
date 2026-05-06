"""
Pulse Lab Hyperspectral Data Processor
======================================

功能：
1. 動態CSV處理（無表頭、列數不固定）
2. 背景訊號標準化（Normalization）
3. Savitzky-Golay濾波（可選）
4. Plotly可視化
5. Excel雙頁式匯出

作者: Nelson @ NTHU Pulse Lab
版本: 1.0
"""

import pandas as pd
import numpy as np
import logging
from pathlib import Path
from typing import Tuple, List, Dict, Optional
from scipy.signal import savgol_filter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# ============================================================
# 日誌設置
# ============================================================

logger = logging.getLogger(__name__)

def setup_logger(log_file: str = "hyperspectral_processing.log"):
    """設置日誌系統"""
    handler = logging.FileHandler(log_file)
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    logger.setLevel(logging.INFO)

# ============================================================
# 核心數據處理函數
# ============================================================

class HyperspectralProcessor:
    """高光譜數據處理類"""
    
    def __init__(self, wavelength_col_index: int = 0):
        """
        初始化處理器
        
        Args:
            wavelength_col_index: 波長列的索引（預設第一列）
        """
        self.wavelength_col_index = wavelength_col_index
        self.df = None
        self.wavelengths = None
        self.measurement_cols = None
        self.background_cols = None
        self.normalized_cols = None
        self.processing_log = []
        
    def read_csv(self, file_path: str, bg_count: int = 3) -> bool:
        """
        讀取CSV檔案（無表頭）
        
        Args:
            file_path: CSV檔案路徑
            bg_count: 背景訊號的列數（從最後往前）
            
        Returns:
            bool: 是否成功讀取
        """
        try:
            # 無表頭讀取
            self.df = pd.read_csv(file_path, header=None)
            
            # 驗證數據有效性
            if self.df.empty:
                logger.error("CSV檔案為空")
                self.processing_log.append("❌ 錯誤：CSV檔案為空")
                return False
            
            if len(self.df.columns) < bg_count + 2:
                logger.error(f"列數不足。至少需要 {bg_count + 2} 列")
                self.processing_log.append(
                    f"❌ 錯誤：列數不足。至少需要 {bg_count + 2} 列，實際：{len(self.df.columns)} 列"
                )
                return False
            
            # 自動命名列
            self._auto_name_columns(bg_count)
            
            logger.info(f"成功讀取 {file_path}，大小：{self.df.shape}")
            self.processing_log.append(
                f"✅ 成功讀取CSV檔案 ({self.df.shape[0]} 個數據點, {self.df.shape[1]} 列)"
            )
            
            return True
        
        except Exception as e:
            logger.error(f"讀取CSV失敗：{str(e)}")
            self.processing_log.append(f"❌ 讀取失敗：{str(e)}")
            return False
    
    def read_from_buffer(self, file_buffer, bg_count: int = 3) -> bool:
        """
        從上傳的檔案緩衝區讀取CSV
        用於Streamlit的uploaded_file物件
        
        Args:
            file_buffer: Streamlit uploaded_file物件
            bg_count: 背景訊號的列數
            
        Returns:
            bool: 是否成功讀取
        """
        try:
            # 重要：重置檔案指標
            file_buffer.seek(0)
            
            # 讀取CSV（無表頭）
            self.df = pd.read_csv(file_buffer, header=None)
            
            # 驗證數據
            if self.df.empty:
                logger.error("上傳檔案為空")
                self.processing_log.append("❌ 錯誤：上傳檔案為空")
                return False
            
            if len(self.df.columns) < bg_count + 2:
                logger.error(f"列數不足。需要 {bg_count + 2} 列，實際 {len(self.df.columns)} 列")
                self.processing_log.append(
                    f"❌ 錯誤：列數不足。需要最少 {bg_count + 2} 列"
                )
                return False
            
            # 自動命名
            self._auto_name_columns(bg_count)
            
            logger.info(f"成功讀取上傳檔案，大小：{self.df.shape}")
            self.processing_log.append(
                f"✅ 成功讀取檔案 ({self.df.shape[0]} 個數據點, {self.df.shape[1]} 列)"
            )
            
            return True
        
        except Exception as e:
            logger.error(f"讀取上傳檔案失敗：{str(e)}")
            self.processing_log.append(f"❌ 讀取失敗：{str(e)}")
            return False
    
    def _auto_name_columns(self, bg_count: int):
        """自動命名列"""
        cols = self.df.columns.tolist()
        
        # 波長列
        self.wavelengths = self.df.iloc[:, self.wavelength_col_index]
        self.df = self.df.rename(columns={
            self.wavelength_col_index: 'Wavelength'
        })
        
        # 背景訊號列（最後N列）
        self.background_cols = [f'Background_{i+1}' for i in range(bg_count)]
        bg_mapping = {cols[-(bg_count-i)]: self.background_cols[i] 
                     for i in range(bg_count)}
        self.df = self.df.rename(columns=bg_mapping)
        
        # 測量點數據列（中間所有列）
        n_data_cols = len(cols) - bg_count - 1
        self.measurement_cols = [f'Measurement_{i+1}' for i in range(n_data_cols)]
        data_mapping = {cols[i+1]: self.measurement_cols[i] 
                       for i in range(n_data_cols)}
        self.df = self.df.rename(columns=data_mapping)
    
    def normalize(self, handle_zero: float = 1e-9) -> bool:
        """
        進行背景訊號標準化
        
        Args:
            handle_zero: 避免除以零的小值
            
        Returns:
            bool: 是否成功
        """
        try:
            if self.df is None:
                logger.error("尚未讀取數據")
                self.processing_log.append("❌ 錯誤：尚未讀取數據")
                return False
            
            # 計算背景平均值
            self.df['Background_AVG'] = self.df[self.background_cols].mean(axis=1)
            
            # 標準化所有測量點
            self.normalized_cols = []
            for col in self.measurement_cols:
                normalized_col = f'{col}_normalized'
                # 防止除以零
                self.df[normalized_col] = self.df[col] / (
                    self.df['Background_AVG'] + handle_zero
                )
                self.normalized_cols.append(normalized_col)
            
            logger.info(f"成功標準化 {len(self.measurement_cols)} 個測量點")
            self.processing_log.append(
                f"✅ 完成標準化 ({len(self.measurement_cols)} 個測量點)"
            )
            
            return True
        
        except Exception as e:
            logger.error(f"標準化失敗：{str(e)}")
            self.processing_log.append(f"❌ 標準化失敗：{str(e)}")
            return False
    
    def apply_savitzky_golay_filter(self, 
                                   window_length: int = 51,
                                   polyorder: int = 3) -> bool:
        """
        應用 Savitzky-Golay 濾波（去噪）
        
        Args:
            window_length: 視窗長度（必須是奇數）
            polyorder: 多項式階數
            
        Returns:
            bool: 是否成功
        """
        try:
            if self.normalized_cols is None:
                logger.error("尚未進行標準化")
                self.processing_log.append("❌ 錯誤：尚未進行標準化")
                return False
            
            # 確保視窗長度是奇數
            if window_length % 2 == 0:
                window_length += 1
            
            # 確保視窗長度不超過數據長度
            if window_length > len(self.df):
                window_length = len(self.df) if len(self.df) % 2 == 1 else len(self.df) - 1
            
            # 應用濾波
            for col in self.normalized_cols:
                filtered_col = f'{col}_filtered'
                self.df[filtered_col] = savgol_filter(
                    self.df[col],
                    window_length=window_length,
                    polyorder=polyorder
                )
            
            logger.info(f"成功應用Savitzky-Golay濾波 (window={window_length})")
            self.processing_log.append(
                f"✅ 應用Savitzky-Golay濾波 (視窗={window_length})"
            )
            
            return True
        
        except Exception as e:
            logger.error(f"濾波失敗：{str(e)}")
            self.processing_log.append(f"❌ 濾波失敗：{str(e)}")
            return False
    
    def export_to_csv(self, output_path: str, include_filtered: bool = False) -> bool:
        """將處理後的數據導出為 CSV (具備自動欄位識別功能)"""
        try:
            if self.df is None:
                return False
            
            # --- 自動識別第一欄 (波長) ---
            all_cols = self.df.columns.tolist()
            first_col = all_cols[0] 
            
            # 決定要導出的欄位清單
            # 包含：波長欄 + 正規化後的測量數據
            cols = [first_col] + self.normalized_cols
            
            # 如果用戶勾選要包含濾波數據
            if include_filtered:
                filtered_cols = [f'{col}_filtered' for col in self.normalized_cols]
                # 確保這些濾波後的列真的存在於數據框中
                actual_filtered = [c for c in filtered_cols if c in self.df.columns]
                cols.extend(actual_filtered)
            
            # 執行導出 (使用 utf-8-sig 確保 Excel 開啟不亂碼)
            self.df[cols].to_csv(output_path, index=False, encoding='utf-8-sig')
            return True
        except Exception as e:
            if hasattr(self, 'logger'):
                self.logger.error(f"❌ CSV 導出失敗: {str(e)}")
            return False

    def _apply_direct_formatting(self, workbook):
        """直接對記憶體中的 Workbook 對象進行美化"""
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for sheetname in workbook.sheetnames:
            ws = workbook[sheetname]
            ws.freeze_panes = 'B2' # 凍結第一列與第一欄
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
    
    def _format_excel(self, file_path: str):
        """美化Excel檔案（自動調整欄寬、凍結第一列等）"""
        try:
            wb = openpyxl.load_workbook(file_path)
            
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                
                # 凍結第一列（波長）
                ws.freeze_panes = 'B1'
                
                # 自動調整欄寬
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # 標頭格式
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", 
                                        fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            wb.save(file_path)
            logger.info("Excel檔案格式化完成")
        
        except Exception as e:
            logger.warning(f"Excel格式化失敗（但檔案已保存）：{str(e)}")
    
    def get_normalized_data(self, use_filtered: bool = False) -> pd.DataFrame:
        """
        獲取標準化數據用於可視化
        
        Args:
            use_filtered: 是否使用濾波後的數據
            
        Returns:
            pd.DataFrame: 包含波長和標準化數據的DataFrame
        """
        if self.normalized_cols is None:
            logger.error("尚未進行標準化")
            return None
        
        cols = ['Wavelength'] + self.normalized_cols
        
        if use_filtered:
            filtered_cols = [f'{col}_filtered' for col in self.normalized_cols]
            cols = ['Wavelength'] + filtered_cols
        
        return self.df[cols].copy()
    
    def get_processing_summary(self) -> Dict:
        """獲取處理摘要"""
        return {
            'timestamp': datetime.now().isoformat(),
            'data_shape': self.df.shape if self.df is not None else None,
            'measurement_points': len(self.measurement_cols) if self.measurement_cols else 0,
            'background_signals': len(self.background_cols) if self.background_cols else 0,
            'processing_steps': self.processing_log
        }

# ============================================================
# 向下相容性：保持原有函數簽名
# ============================================================

def process_hyperspectral_data(df: pd.DataFrame, 
                              bg_count: int = 3) -> Tuple[pd.DataFrame, List, List]:
    """
    原有函數（向下相容）
    
    Args:
        df: 傳入的原始 DataFrame (第一列為波長)
        bg_count: 背景訊號的數量 (從最後一列往前推)
        
    Returns:
        Tuple: (處理後的DataFrame, 測量列表, 背景列表)
    """
    processor = HyperspectralProcessor()
    
    # 手動設置數據
    processor.df = df.copy()
    processor._auto_name_columns(bg_count)
    processor.normalize()
    
    return processor.df, processor.measurement_cols, processor.background_cols

# ============================================================
# 初始化
# ============================================================

if __name__ == "__main__":
    setup_logger()
    print("✅ 高光譜處理模組已載入")